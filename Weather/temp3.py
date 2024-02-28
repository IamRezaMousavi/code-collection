# -*- coding: utf-8 -*-
"""
Created on Fri May 28 03:29:55 2021

@author: Mohammad
"""

import requests, json
from pprint import pprint as pp
import pprint
from xlwt import Workbook
import matplotlib.pyplot as plt
import openpyxl

#url_base = "http://api.openweathermap.org/data/2.5/weather?q="
url_base = "http://api.openweathermap.org/data/2.5/forecast?q="
#cityname = input("Enter city name:").title()
cityname = "Abadan"
APIKey = "58339b029b4279319dfd339d5f21d532"
url = url_base + cityname + "&appid=" + APIKey #+ "&lang=fa"

try:
    print("Try to send a request")
    data = requests.get(url)
    data = data.json()
    data = json.dumps(data)
    print("Request is work seccessful")

    with open(cityname+"Best.txt", "a+") as f:
        f.write(data)
        f.write("\n")
        print("Write is done.")
    
    data = json.loads(data)

except:
    print("Request is not working")
    try:
        print("Try to load from database")
        with open(cityname+"Best.txt") as r:
            olddata = r.readlines()
            print("Read is Seccessful.")
    
        data = json.loads(olddata[-1])
        pp(data)
    except:
        print("Database is not found")
else:
    pp(data)

# Edit City Name In Data Dic
data["city"]["name"] = cityname

# Save Pretty Data
print()
prettydata = pprint.pformat(data)
with open(cityname + "PrettyDataBest.txt", "a+") as pd:
    pd.write("\n")
    pd.write(prettydata)
print(cityname + " Pretty Data was written")

# Print Temp Values
print()
now = data["list"][0]["dt_txt"]
far = data["list"][0]["main"]["temp"]
print("Temp of ", now," is ", far, " f")
c = data["list"][0]["main"]["temp"] - 273.15
print("Temp of ", now," is ", round(c, 2), " c")
print()

# Creat Data List
temp = [data["list"][i]["main"]["temp"] for i in range(len(data["list"]))]
temptime = [int((data["list"][i]["dt"]-data["list"][0]["dt"])/3600) for i in range(len(data["list"]))]

# Plot Temp Data
#plt.style.use("seaborn-dark-palette")
plt.plot(temptime, [t-273.15 for t in temp], color= "#FF0000")
plt.grid()
plt.xlabel("Time (h)")
plt.ylabel("Temp (c)")
plt.title(cityname + " Temp Data\n from " + now)
plt.savefig(cityname + ".png", format="png", dpi=1000, transparent=False)
print("The plot is saved in png file")
print()


# Write Data in Xls File
wb = Workbook()
tempsheet = wb.add_sheet("TempData")
for i in range(len(data["list"])):
    tempsheet.write(i,0,temp[i])
    tempsheet.write(i,1,temptime[i])
    tempsheet.write(i,2,data["list"][i]["dt"])
wb.save(cityname + "TempDataBestBest.xls")
print("The temp data is saved in xls file")
print()

# Write Data in Xlsx File
book = openpyxl.Workbook()
book.create_sheet("Temp")
sheet = book.get_sheet_by_name("Temp")
for i in range(1, len(data["list"])+1):
    sheet.cell(row=i,column=1).value = temp[i-1]
    sheet.cell(row=i,column=2).value = temptime[i-1]
    sheet.cell(row=i,column=3).value = data["list"][i-1]["dt"]

extrasheet = book.get_sheet_by_name("Sheet")
book.remove(extrasheet)
book.save(cityname + "TempDataBestBest.xlsx")
